import streamlit as st
import pandas as pd
import difflib
import os
from openpyxl import load_workbook

st.set_page_config(page_title="Match de Vagas", page_icon="💼", layout="centered")

CAMINHO_VAGAS = "Streamlit_desafio_5/Vagas.xlsx"
CAMINHO_CANDIDATOS = "Streamlit_desafio_5/Modelo_Candidato_Simplificado.xlsx"

@st.cache_data
def carregar_vagas():
    vagas = pd.read_excel(CAMINHO_VAGAS)
    vagas.rename(columns={
        "Título da vaga": "Vaga",
        "Tipo": "Tipo",
        "Área": "Area",
        "Habilidades": "Habilidades",
        "Nível de inglês": "Level de Ingles",
        "Nível de espanhol": "Level de Espanhol",
        "Outros idiomas": "Outros idiomas",
        "Disponível para viagens": "Precisa Viajar",
        "Possui equipamento próprio": "Precisa de Equipamento",
        "Empresa": "Empresa",
        "Descrição da vaga": "Descricao",
        "Salário pago": "Salario"
    }, inplace=True)
    return vagas

def calcular_score(candidato, vaga):
    score = 0
    peso_total = 0

    if isinstance(candidato["Título da vaga desejada"], str) and isinstance(vaga["Vaga"], str):
        ratio = difflib.SequenceMatcher(None, candidato["Título da vaga desejada"].lower(), vaga["Vaga"].lower()).ratio()
        score += ratio * 2
        peso_total += 2

    if isinstance(candidato["Tipo da vaga desejada"], str) and isinstance(vaga["Tipo"], str):
        if candidato["Tipo da vaga desejada"].lower() == vaga["Tipo"].lower():
            score += 1
        peso_total += 1

    if isinstance(candidato["Área de interesse"], str) and isinstance(vaga["Area"], str):
        if candidato["Área de interesse"].lower() in vaga["Area"].lower():
            score += 1
        peso_total += 1

    if isinstance(candidato["Nível de inglês"], str) and isinstance(vaga["Level de Ingles"], str):
        if candidato["Nível de inglês"].lower() in vaga["Level de Ingles"].lower():
            score += 1
        peso_total += 1

    if isinstance(candidato["Nível de espanhol"], str) and isinstance(vaga["Level de Espanhol"], str):
        if candidato["Nível de espanhol"].lower() in vaga["Level de Espanhol"].lower():
            score += 1
        peso_total += 1

    if isinstance(candidato["Possui equipamento próprio? (Sim/Não)"], str) and isinstance(vaga["Precisa de Equipamento"], str):
        if candidato["Possui equipamento próprio? (Sim/Não)"].lower() == "sim" and "não" not in vaga["Precisa de Equipamento"].lower():
            score += 1
        peso_total += 1

    if isinstance(candidato["Disponível para viagens? (Sim/Não)"], str) and isinstance(vaga["Precisa Viajar"], str):
        if candidato["Disponível para viagens? (Sim/Não)"].lower() == vaga["Precisa Viajar"].lower():
            score += 1
        peso_total += 1

    candidato_skills = set(h.lower() for h in candidato["Competências técnicas"]) if isinstance(candidato["Competências técnicas"], list) else set()
    vaga_skills = set(str(vaga["Habilidades"]).lower().split(",")) if isinstance(vaga["Habilidades"], str) else set()
    vaga_skills = set(h.strip() for h in vaga_skills)
    intersecao = candidato_skills.intersection(vaga_skills)
    if vaga_skills:
        score += len(intersecao) / len(vaga_skills) * 3
        peso_total += 3

    if isinstance(candidato["Outros idiomas"], str) and isinstance(vaga["Outros idiomas"], str):
        if candidato["Outros idiomas"].lower() in vaga["Outros idiomas"].lower():
            score += 1
        peso_total += 1

    if isinstance(vaga["Descricao"], str) and isinstance(candidato["Competências técnicas"], list):
        desc_text = vaga["Descricao"].lower()
        match_count = sum(1 for skill in candidato["Competências técnicas"] if skill.lower() in desc_text)
        if len(candidato["Competências técnicas"]) > 0:
            score += (match_count / len(candidato["Competências técnicas"])) * 2
            peso_total += 2

    return round((score / peso_total) * 100, 2) if peso_total else 0

# ---------- INTERFACE ---------- #
vagas_df = carregar_vagas()
titulos_disponiveis = sorted(vagas_df["Vaga"].dropna().unique())
areas_disponiveis = sorted(vagas_df["Area"].dropna().unique())

todas_habilidades = vagas_df["Habilidades"].dropna().str.cat(sep=",").lower().split(",")
habilidades_unicas = sorted(set(h.strip().capitalize() for h in todas_habilidades if h.strip() != ""))

st.title("🔍 Plataforma de Match de Vagas")
st.markdown("Preencha abaixo e veja quais vagas combinam com você!")

with st.form("formulario_candidato"):
    st.subheader("📄 Dados do Candidato")
    nome = st.text_input("Nome completo")
    cpf = st.text_input("CPF")
    cidade = st.text_input("Cidade onde mora")
    titulo = st.selectbox("Título da vaga desejada", titulos_disponiveis)
    tipo = st.selectbox("Tipo da vaga desejada", ["Júnior", "Pleno", "Sênior"])
    area = st.selectbox("Área de interesse", areas_disponiveis)
    ingles = st.selectbox("Nível de inglês", ["Nenhum", "Básico", "Intermediário", "Avançado", "Fluente"])
    espanhol = st.selectbox("Nível de espanhol", ["Nenhum", "Básico", "Intermediário", "Avançado", "Fluente"])
    outros_idiomas = st.text_input("Outros idiomas")
    tecnicas = st.multiselect("Competências técnicas", habilidades_unicas)
    comportamentais = st.text_area("Competências comportamentais")
    viagens = st.selectbox("Disponível para viagens?", ["Sim", "Não"])
    equipamento = st.selectbox("Possui equipamento próprio?", ["Sim", "Não"])
    salario = st.text_input("Expectativa salarial")

    enviado = st.form_submit_button("🔎 Encontrar Vagas")

if enviado:
    candidato = {
        "Nome completo": nome,
        "CPF": cpf,
        "Cidade": cidade,
        "Título da vaga desejada": titulo,
        "Tipo da vaga desejada": tipo,
        "Área de interesse": area,
        "Nível de inglês": ingles,
        "Nível de espanhol": espanhol,
        "Outros idiomas": outros_idiomas,
        "Competências técnicas": tecnicas,
        "Competências comportamentais": comportamentais,
        "Disponível para viagens? (Sim/Não)": viagens,
        "Possui equipamento próprio? (Sim/Não)": equipamento,
        "Expectativa salarial": salario
    }

    # ---------- SALVAR NO EXCEL ----------
    novo_candidato_df = pd.DataFrame([candidato])
    novo_candidato_df["Competências técnicas"] = [", ".join(tecnicas)]

    if os.path.exists(CAMINHO_CANDIDATOS):
        with pd.ExcelWriter(CAMINHO_CANDIDATOS, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            book = load_workbook(CAMINHO_CANDIDATOS)
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            startrow = writer.sheets['Sheet1'].max_row
            novo_candidato_df.to_excel(writer, index=False, header=False, startrow=startrow)
    else:
        novo_candidato_df.to_excel(CAMINHO_CANDIDATOS, index=False)

    st.success("📝 Seus dados foram salvos com sucesso!")

    # ---------- CÁLCULO DE MATCH ----------
    st.info("🔄 Processando suas informações...")
    vagas_df["ia_score"] = vagas_df.apply(lambda row: calcular_score(candidato, row), axis=1)
    top_vagas = vagas_df.sort_values(by="ia_score", ascending=False).head(5)

    st.success("✅ Veja abaixo suas vagas com maior compatibilidade!")
    st.subheader("🏆 Top 5 Vagas Compatíveis")

    for _, vaga in top_vagas.iterrows():
        st.markdown(f"### {vaga['Vaga']}")
        st.markdown(f"**Tipo:** {vaga['Tipo']}")
        st.markdown(f"**Área:** {vaga['Area']}")
        st.markdown(f"**Cliente:** {vaga['Empresa']}")
        st.markdown(f"**Score de compatibilidade:** {vaga['ia_score']}%")
        st.markdown(f"**Requisitos:** {vaga['Habilidades']}")
        st.markdown(f"**Descrição:** {vaga['Descricao']}")
        st.markdown(f"**Salário oferecido:** {vaga['Salario']}")
        st.markdown("---")
